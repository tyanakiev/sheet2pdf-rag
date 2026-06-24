import os
import sys
import shutil
import tempfile
import logging
from pathlib import Path
from dataclasses import dataclass
from typing import List, Optional, Literal

import openpyxl

logger = logging.getLogger(__name__)

@dataclass
class ConversionResult:
    original_excel_path: str
    pdf_path: str
    sheet_name: Optional[str]
    engine_used: str

class ExcelConverter:
    """
    Core engine to convert Excel files to PDF.
    Auto-detects the operating system and uses the best available engine.
    """
    
    def __init__(self, engine: Literal["auto", "win32", "libreoffice"] = "auto"):
        self.engine_choice = engine
        self._engine = self._resolve_engine()

    def _resolve_engine(self) -> str:
        if self.engine_choice != "auto":
            return self.engine_choice

        if sys.platform == "win32":
            try:
                import win32com.client
                # Basic check to see if Excel is installed
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Quit()
                return "win32"
            except ImportError:
                logger.warning("pywin32 not installed. Falling back to LibreOffice.")
            except Exception as e:
                logger.warning(f"Could not dispatch Excel COM object: {e}. Falling back to LibreOffice.")
        
        # Check if soffice is in path
        if shutil.which("soffice") or shutil.which("libreoffice"):
            return "libreoffice"
        
        raise RuntimeError(
            "No suitable Excel-to-PDF engine found. "
            "Please install Microsoft Excel (Windows) or LibreOffice (Cross-platform)."
        )

    def convert(self, excel_path: str | Path, output_dir: str | Path, split_sheets: bool = True) -> List[ConversionResult]:
        excel_path = Path(excel_path).resolve()
        output_dir = Path(output_dir).resolve()
        output_dir.mkdir(parents=True, exist_ok=True)

        if not excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        if self._engine == "win32":
            return self._convert_win32(excel_path, output_dir, split_sheets)
        elif self._engine == "libreoffice":
            return self._convert_libreoffice(excel_path, output_dir, split_sheets)
        else:
            raise ValueError(f"Unknown engine: {self._engine}")

    def _convert_win32(self, excel_path: Path, output_dir: Path, split_sheets: bool) -> List[ConversionResult]:
        import win32com.client
        
        results = []
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        try:
            wb = excel.Workbooks.Open(str(excel_path))
            base_name = excel_path.stem
            
            if split_sheets:
                for sheet in wb.Worksheets:
                    pdf_filename = f"{base_name}__{sheet.Name}.pdf"
                    pdf_path = output_dir / pdf_filename
                    
                    # 0 corresponds to xlTypePDF
                    sheet.ExportAsFixedFormat(0, str(pdf_path))
                    
                    results.append(ConversionResult(
                        original_excel_path=str(excel_path),
                        pdf_path=str(pdf_path),
                        sheet_name=sheet.Name,
                        engine_used="win32"
                    ))
            else:
                pdf_filename = f"{base_name}.pdf"
                pdf_path = output_dir / pdf_filename
                wb.ExportAsFixedFormat(0, str(pdf_path))
                results.append(ConversionResult(
                    original_excel_path=str(excel_path),
                    pdf_path=str(pdf_path),
                    sheet_name=None,
                    engine_used="win32"
                ))
        finally:
            # Ensure workbook is closed and Excel application quits
            try:
                wb.Close(SaveChanges=False)
            except:
                pass
            excel.Quit()
            
        return results

    def _convert_libreoffice(self, excel_path: Path, output_dir: Path, split_sheets: bool) -> List[ConversionResult]:
        import subprocess
        
        soffice_path = shutil.which("soffice") or shutil.which("libreoffice")
        if not soffice_path:
            raise RuntimeError("LibreOffice 'soffice' executable not found in PATH.")
            
        results = []
        base_name = excel_path.stem
        
        if split_sheets:
            # Load workbook using openpyxl to split sheets
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            sheet_names = wb.sheetnames
            
            # For each sheet, we create a temporary workbook with only that sheet
            with tempfile.TemporaryDirectory() as temp_dir:
                temp_dir_path = Path(temp_dir)
                
                for sheet_name in sheet_names:
                    # We have to reload to not destroy the original workbook state
                    temp_wb = openpyxl.load_workbook(excel_path)
                    for name in temp_wb.sheetnames:
                        if name != sheet_name:
                            del temp_wb[name]
                    
                    temp_xlsx_path = temp_dir_path / f"temp_{sheet_name}.xlsx"
                    temp_wb.save(temp_xlsx_path)
                    
                    # Run libreoffice headless conversion on this single-sheet workbook
                    cmd = [
                        soffice_path,
                        "--headless",
                        "--convert-to", "pdf",
                        "--outdir", str(output_dir),
                        str(temp_xlsx_path)
                    ]
                    subprocess.run(cmd, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                    
                    # The output PDF will be named temp_{sheet_name}.pdf
                    temp_pdf_path = output_dir / f"temp_{sheet_name}.pdf"
                    final_pdf_path = output_dir / f"{base_name}__{sheet_name}.pdf"
                    
                    if temp_pdf_path.exists():
                        temp_pdf_path.rename(final_pdf_path)
                        results.append(ConversionResult(
                            original_excel_path=str(excel_path),
                            pdf_path=str(final_pdf_path),
                            sheet_name=sheet_name,
                            engine_used="libreoffice"
                        ))
        else:
            cmd = [
                soffice_path,
                "--headless",
                "--convert-to", "pdf",
                "--outdir", str(output_dir),
                str(excel_path)
            ]
            subprocess.run(cmd, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            
            expected_pdf = output_dir / f"{base_name}.pdf"
            if expected_pdf.exists():
                results.append(ConversionResult(
                    original_excel_path=str(excel_path),
                    pdf_path=str(expected_pdf),
                    sheet_name=None,
                    engine_used="libreoffice"
                ))
                
        return results
